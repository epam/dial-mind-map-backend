import asyncio
import json
import re
from datetime import date
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

from dial_api import get_graph


async def load_graphs(app_ids: List[str]) -> Dict[str, Any]:
    tasks = [get_graph(app_id) for app_id in app_ids]
    results = await asyncio.gather(*tasks)
    return {app_id: graph for app_id, graph in zip(app_ids, results)}


def calculate_number_words(a: str) -> int:
    return len(re.findall(r"[A-Za-z]+", a))


def fit_columns(ws: Worksheet):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2


def build_border(ws: Worksheet, x1: int, y1: int, x2: int, y2: int):
    medium = Side(border_style="medium", color="000000")

    if x1 != x2 and y1 != y2:
        for x in range(x1, x2 + 1):
            ws.cell(row=x, column=y1).border = Border(left=medium)
            ws.cell(row=x, column=y2).border = Border(right=medium)

        for y in range(y1, y2 + 1):
            ws.cell(row=x1, column=y).border = Border(top=medium)
            ws.cell(row=x2, column=y).border = Border(bottom=medium)

        ws.cell(row=x1, column=y1).border = Border(left=medium, top=medium)
        ws.cell(row=x1, column=y2).border = Border(right=medium, top=medium)
        ws.cell(row=x2, column=y1).border = Border(left=medium, bottom=medium)
        ws.cell(row=x2, column=y2).border = Border(right=medium, bottom=medium)
    elif x1 == x2:
        for y in range(y1 + 1, y2):
            ws.cell(row=x1, column=y).border = Border(top=medium, bottom=medium)

        ws.cell(row=x1, column=y1).border = Border(
            left=medium, top=medium, bottom=medium
        )
        ws.cell(row=x1, column=y2).border = Border(
            right=medium, top=medium, bottom=medium
        )
    elif y1 == y2:
        for x in range(x1 + 1, x2):
            ws.cell(row=x, column=y1).border = Border(left=medium, right=medium)

        ws.cell(row=x1, column=y1).border = Border(
            left=medium, top=medium, right=medium
        )
        ws.cell(row=x2, column=y1).border = Border(
            left=medium, bottom=medium, right=medium
        )
    else:
        ws.cell(row=x1, column=y1).border = Border(
            left=medium, bottom=medium, right=medium, top=medium
        )


def number_runs(eval: Any) -> int:
    for _, results in enumerate(eval["results"]):
        unique_graders = set()

        for result in results:
            unique_graders.add(result["grader_name"])

        return len(results) // len(unique_graders)


def build_graders_sheet(ws: Worksheet, eval: Any):
    ws["A1"] = "Name"
    ws["A1"].font = Font(bold=True)
    ws["B1"] = "Type"
    ws["B1"].font = Font(bold=True)
    ws["C1"] = "Description"
    ws["C1"].font = Font(bold=True)
    ws["D1"] = "Prompt"
    ws["D1"].font = Font(bold=True)
    ws["E1"] = "Metadata"
    ws["E1"].font = Font(bold=True)
    for i, grader in enumerate(eval["graders"]):
        ws.cell(row=i + 2, column=1, value=grader["name"])
        ws.cell(row=i + 2, column=2, value=grader["type"])
        ws.cell(row=i + 2, column=3, value=grader.get("description", ""))
        ws.cell(
            row=i + 2,
            column=4,
            value=grader["prompt"] if grader["type"] == "Graph" else "",
        )
        ws.cell(row=i + 2, column=5, value=json.dumps(grader))

    fit_columns(ws)

    build_border(ws, 1, 1, 1 + len(eval["graders"]), 5)
    build_border(ws, 1, 1, 1, 5)


def build_overview_sheet(ws: Worksheet, eval: Any, graphs: Dict[str, Any]):
    ws["A1"] = "Evaluation Metadata"
    ws["A1"].font = Font(bold=True)

    ws["A2"] = "Date"
    ws["A2"].font = Font(bold=True)
    ws["B2"] = date.today().strftime("%Y-%m-%d")

    ws["A3"] = "LLM model name and version"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = "Manual"

    ws["A4"] = "Number of publications"
    ws["A4"].font = Font(bold=True)
    ws["B4"] = len(eval["results"])

    ws["A5"] = "Number of runs per publication"
    ws["A5"].font = Font(bold=True)
    ws["B5"] = number_runs(eval)

    ws["A6"] = "Number of graders"
    ws["A6"].font = Font(bold=True)
    ws["B6"] = len(eval["graders"])

    ws["A9"] = "Publication name"
    ws["A9"].font = Font(bold=True)

    ws["B9"] = "Avg score"
    ws["B9"].font = Font(bold=True)

    ws["C9"] = "Avg nodes"
    ws["C9"].font = Font(bold=True)

    ws["D9"] = "Avg words"
    ws["D9"].font = Font(bold=True)

    ws["E9"] = "Min words"
    ws["E9"].font = Font(bold=True)

    ws["F9"] = "Max words"
    ws["F9"].font = Font(bold=True)

    ws.cell(
        row=10 + len(eval["results"]),
        column=6,
        value="Avg per grader",
    )
    for i, grader in enumerate(eval["graders"]):
        grader_title = ws.cell(row=9, column=7 + i, value=grader["name"])
        grader_title.font = Font(bold=True)

        common_passed_by_grader = 0
        common_total_by_grader = 0
        for j, results in enumerate(eval["results"]):
            passed_by_grader = 0
            total_by_grader = 0

            for result in results:
                if result["grader_name"] == grader["name"]:
                    if result["result"]["result"] == "pass":
                        passed_by_grader += 1
                    total_by_grader += 1

            if total_by_grader:
                ws.cell(
                    row=10 + j,
                    column=7 + i,
                    value=round(passed_by_grader / total_by_grader, 2),
                )

            common_passed_by_grader += passed_by_grader
            common_total_by_grader += total_by_grader

        ws.cell(
            row=10 + len(eval["results"]),
            column=7 + i,
            value=round(common_passed_by_grader / common_total_by_grader, 2),
        )

    total = 0
    passed = 0
    used_sheets = {}
    for i, results in enumerate(eval["results"]):
        run_name = results[0]["run_name"]

        total_by_run = 0
        passed_by_run = 0
        number_nodes = 0
        number_words = 0
        min_words = 10**9
        max_words = 0
        for result in results:
            app_id = result["app_id"]
            graph = graphs[app_id]

            if result["result"]["result"] == "pass":
                passed_by_run += 1
            total_by_run += 1

            for node in graph["nodes"]:
                node_number_words = calculate_number_words(node["data"]["details"])

                number_words += node_number_words
                min_words = min(min_words, node_number_words)
                max_words = max(max_words, node_number_words)
                number_nodes += 1

        sheet_name = run_name[:30]

        if sheet_name in used_sheets:
            original_sheet_name = sheet_name
            sheet_name = f"{sheet_name}_{used_sheets[original_sheet_name] + 1}"
            used_sheets[original_sheet_name] += 1
        else:
            used_sheets[sheet_name] = 0

        run_name_cell = ws.cell(row=10 + i, column=1)
        run_name_cell.value = f'=HYPERLINK("#\'{sheet_name}\'!A1", "{run_name}")'
        run_name_cell.style = "Hyperlink"

        ws.cell(row=10 + i, column=2, value=round(passed_by_run / total_by_run, 2))
        ws.cell(row=10 + i, column=3, value=round(number_nodes / len(results), 2))
        if number_nodes:
            ws.cell(row=10 + i, column=4, value=round(number_words / number_nodes, 2))
            ws.cell(row=10 + i, column=5, value=min_words)
            ws.cell(row=10 + i, column=6, value=max_words)

        total += total_by_run
        passed += passed_by_run

    ws["A7"] = "Overall score"
    ws["A7"].font = Font(bold=True)
    ws["B7"] = round(passed / total, 2)

    fit_columns(ws)

    build_border(ws, 1, 1, 7, 2)

    build_border(ws, 9, 1, 9 + len(eval["results"]), 6 + len(eval["graders"]))
    build_border(ws, 9, 1, 9, 6 + len(eval["graders"]))
    build_border(ws, 9, 1, 9 + len(eval["results"]), 1)
    build_border(ws, 9, 1, 9, 1)

    build_border(
        ws,
        10 + len(eval["results"]),
        6,
        10 + len(eval["results"]),
        6 + len(eval["graders"]),
    )


def build_result_sheet(ws: Worksheet, eval: Any, results: Any, graphs: Dict[str, Any]):
    max_rows = 0

    run_title = ws.cell(row=1, column=1, value="Run ID")
    run_title.font = Font(bold=True)

    for i, grader in enumerate(eval["graders"]):
        grader_title = ws.cell(row=1, column=2 + i, value=grader["name"])
        grader_title.font = Font(bold=True)

        passed_by_grader = 0
        total_by_grader = 0

        current_row = 2
        for result in results:
            if result["grader_name"] == grader["name"]:
                cell = ws.cell(
                    row=current_row,
                    column=2 + i,
                    value=1 if result["result"]["result"] == "pass" else 0,
                )
                cell.comment = Comment(
                    text=result["result"]["reasoning"], author="System"
                )
                cell.alignment = Alignment(horizontal="left")

                if result["result"]["result"] == "pass":
                    passed_by_grader += 1
                total_by_grader += 1

                current_row += 1

        max_rows = max(max_rows, current_row)

        if total_by_grader:
            ws.cell(
                row=current_row,
                column=2 + i,
                value=round(passed_by_grader / total_by_grader, 2),
            )

    avg_per_grader = ws.cell(row=max_rows, column=1, value="Avg per grader")
    avg_per_grader.font = Font(bold=True)

    for i in range(2, max_rows):
        ws.cell(row=i, column=1, value=f"#{i - 1}")

    fit_columns(ws)

    build_border(ws, 1, 1, max_rows - 1, 1 + len(eval["graders"]))
    build_border(ws, 1, 1, 1, 1 + len(eval["graders"]))
    build_border(ws, max_rows, 1, max_rows, 1 + len(eval["graders"]))


def prepare_report(eval: Any) -> bytes:
    all_app_ids = {
        result["app_id"]
        for run_result in eval["results"]
        if type(run_result) == list
        for result in run_result
    }

    graphs = asyncio.run(load_graphs(all_app_ids))

    wb = Workbook()
    wb.remove(wb.active)

    build_overview_sheet(wb.create_sheet("Overview"), eval, graphs)
    build_graders_sheet(wb.create_sheet("Graders"), eval)

    for result in eval["results"]:
        run_name = result[0]["run_name"]

        base_name = run_name[:30]

        sheet_name = base_name
        counter = 0

        while sheet_name in wb.sheetnames:
            sheet_name = f"{base_name}{counter}"
            counter += 1

        ws = wb.create_sheet(sheet_name)

        build_result_sheet(ws, eval, result, graphs)

    buffer = BytesIO()
    wb.save(buffer)

    return buffer.getvalue()
